"""
XLSX extractor — handles two distinct sheet layouts:

  1. SIMPLE TABLE sheets — one header row, then data rows below it.
     (e.g. a plain list of wells, a sensor log)

  2. REPORT-STYLE sheets — mixed content: title rows, section headings,
     key-value pairs, AND real tables, all in the same sheet.
     (e.g. "Final Well Status Report" — numbered sections like
     "1. WELL DISPOSITION" each containing either key-value pairs
     or a proper table)

Report-style sheets need ROW-BY-ROW classification, since there's no
single "header row" that applies to the whole sheet — every section
can have its own structure.
"""

import pandas as pd
import openpyxl
from pathlib import Path


def clean_dataframe(df):
    df = df.map(lambda x: x.strip() if isinstance(x, str) else x)
    df = df.replace(r'^\s*$', float('nan'), regex=True)
    df = df.dropna(how='all')
    df = df.reset_index(drop=True)
    return df


def get_used_column_range(ws):
    """
    Returns (min_col, max_col) of columns that actually contain data
    anywhere in the sheet. Some report-style sheets have an entirely
    empty column A (or trailing empty columns), which makes ws.max_column
    misleading for width comparisons.
    """
    used_cols = set()
    for row in ws.iter_rows():
        for cell in row:
            if cell.value is not None and str(cell.value).strip():
                used_cols.add(cell.column)
    if not used_cols:
        return 1, ws.max_column
    return min(used_cols), max(used_cols)


def get_merge_for_cell(ws, row_num, col_num):
    """Returns the merge range that (row_num, col_num) belongs to, or None."""
    for r in ws.merged_cells.ranges:
        if r.min_row <= row_num <= r.max_row and r.min_col <= col_num <= r.max_col:
            return r
    return None


def count_distinct_original_values(ws, row_num):
    """
    Counts how many DISTINCT non-empty values exist in a row, using the
    ORIGINAL (pre-merge-expansion) cell values — only the top-left cell
    of each merge holds a real value; the rest are None until expanded.

    A title row has exactly 1 distinct value (one heading, possibly
    merged across many columns). A key-value row has 2+ distinct values
    (e.g. 'Well Name:' and 'DSEC-AD-1X' are two separate merged cells).
    """
    seen = set()
    for c in range(1, ws.max_column + 1):
        v = ws.cell(row_num, c).value
        if v is not None and str(v).strip():
            seen.add(str(v).strip())
    return len(seen)


def is_full_width_merge(ws, row_num, used_min_col, used_max_col):
    """
    True if row_num has a merge that spans MOST of the sheet's actually-used
    width AND the merge contains only ONE real value (a title/heading has
    a single piece of text, not a key+value pair).

    Checking the value count (not just merge width) is necessary because
    some sheets have merges that extend past the actual data footprint
    (e.g. a key-value merge B4:D4 in a sheet where only columns A-B ever
    hold real data) — comparing widths alone would misclassify those
    key-value rows as titles.
    """
    used_width = used_max_col - used_min_col + 1
    for r in ws.merged_cells.ranges:
        if r.min_row == row_num == r.max_row:
            merge_width = r.max_col - r.min_col + 1
            if merge_width >= used_width - 1:
                return True
    return False


def expand_merged_cells(ws):
    """Build a (row, col) -> value map with merged cells expanded."""
    val_map = {}
    for r in range(1, ws.max_row + 1):
        for c in range(1, ws.max_column + 1):
            val_map[(r, c)] = ws.cell(r, c).value

    for merged in ws.merged_cells.ranges:
        top_val = ws.cell(merged.min_row, merged.min_col).value
        for r in range(merged.min_row, merged.max_row + 1):
            for c in range(merged.min_col, merged.max_col + 1):
                val_map[(r, c)] = top_val

    return val_map


def get_row_values(val_map, row_num, max_col):
    """Returns the row as a list, stripped, with None for empty cells."""
    vals = []
    for c in range(1, max_col + 1):
        v = val_map.get((row_num, c))
        vals.append(str(v).strip() if v is not None and str(v).strip() else None)
    return vals


def dedupe_consecutive(values):
    """
    Removes consecutive duplicate values caused by merge-cell expansion.
    e.g. ['Well Name:', 'DSEC-AD-1X', 'DSEC-AD-1X', 'DSEC-AD-1X']
      -> ['Well Name:', 'DSEC-AD-1X']
    Only collapses ADJACENT duplicates — a genuinely repeated value
    elsewhere in the row (rare) is left alone.
    """
    result = []
    for v in values:
        if not result or v != result[-1]:
            result.append(v)
    return result


def classify_row(row_values, is_merged_title):
    """
    Classifies one row as:
      'blank'        — entirely empty
      'title'        — single value, spans whole merged width (page title or section heading)
      'key_value'    — alternating Key: / Value pairs
      'table_row'    — anything else with multiple values (likely part of a real table)

    Doesn't assume column A holds the first value — some sheets have an
    entirely empty leading column, so we find the first non-empty cell
    and check the key-value pattern starting from there. Also deduplicates
    consecutive identical values from wide merge-cell expansion before
    pattern matching, since otherwise a single Key/Value pair with a
    wide-merged value looks like a 4+ column table row.
    """
    non_empty = dedupe_consecutive([v for v in row_values if v is not None])

    if not non_empty:
        return 'blank'

    if is_merged_title:
        return 'title'

    # key_value pattern: first real value ends with ':', and (if a third
    # real value exists) it also ends with ':' — i.e. alternating Key/Value
    if non_empty[0].rstrip().endswith(':'):
        if len(non_empty) < 3 or non_empty[2].rstrip().endswith(':'):
            return 'key_value'

    return 'table_row'


def render_key_value_row(row_values) -> str:
    """['Well Name:', 'DSEC-AD-1X', 'API Number:', '42-407-23207'] -> markdown line.
    Works on non-empty, deduplicated values only."""
    non_empty = dedupe_consecutive([v for v in row_values if v is not None])
    parts = []
    i = 0
    while i < len(non_empty):
        key = non_empty[i]
        val = non_empty[i + 1] if i + 1 < len(non_empty) else None
        if key and val:
            parts.append(f"**{key}** {val}")
        i += 2
    return '  '.join(parts)


def render_table_block(rows: list[list]) -> str:
    """Converts a list of table rows (first = header) into a markdown table.
    Drops any column that is empty across ALL rows (e.g. an unused leading column A)."""
    rows = [r for r in rows if any(c is not None for c in r)]
    if not rows:
        return ''

    col_n = max(len(r) for r in rows)
    pad = lambda r: [c if c is not None else '' for c in r] + [''] * (col_n - len(r))
    padded = [pad(r) for r in rows]

    # Find columns that have a value in at least one row
    keep_cols = [i for i in range(col_n) if any(r[i] for r in padded)]
    if not keep_cols:
        return ''

    trimmed = [[r[i] for i in keep_cols] for r in padded]
    header, body = trimmed[0], trimmed[1:]

    lines = ['| ' + ' | '.join(header) + ' |',
             '| ' + ' | '.join(['---'] * len(keep_cols)) + ' |']
    for r in body:
        lines.append('| ' + ' | '.join(r) + ' |')
    return '\n'.join(lines)


def sheet_to_markdown(ws) -> str:
    """
    Row-by-row classification and rendering. Handles BOTH simple table
    sheets and report-style mixed-content sheets correctly, since every
    row is judged independently rather than assuming one global header row.
    """
    max_col = ws.max_column
    val_map = expand_merged_cells(ws)
    used_min_col, used_max_col = get_used_column_range(ws)

    parts = []
    table_buffer = []   # accumulates consecutive 'table_row' rows

    def flush_table_buffer():
        if table_buffer:
            md = render_table_block(table_buffer)
            if md:
                parts.append(md)
            table_buffer.clear()

    skip_rows = set()   # rows already consumed as part of a multi-row merge

    for r in range(1, ws.max_row + 1):
        if r in skip_rows:
            continue

        row_values = get_row_values(val_map, r, max_col)
        is_title = (is_full_width_merge(ws, r, used_min_col, used_max_col)
                    and count_distinct_original_values(ws, r) <= 1)

        wide_merge = None
        for mr in ws.merged_cells.ranges:
            if mr.min_row == r and (mr.max_col - mr.min_col + 1) >= 3:
                wide_merge = mr
                break

        # A wide merge is a STANDALONE PARAGRAPH only if there's no other
        # distinct value earlier in the row (e.g. no 'Key:' label before it).
        # If there IS a preceding value, this is really a key-value row
        # where the value happens to be merged wide — let classify_row
        # handle it normally instead of swallowing just the merged value.
        has_preceding_value = False
        if wide_merge is not None:
            for c in range(1, wide_merge.min_col):
                v = ws.cell(r, c).value
                if v is not None and str(v).strip():
                    has_preceding_value = True
                    break

        if wide_merge is not None and not is_title and not has_preceding_value:
            flush_table_buffer()
            val = val_map.get((r, wide_merge.min_col))
            if val and str(val).strip():
                parts.append(str(val).strip())
            # Mark any additional rows this merge spans as already handled
            for skip_r in range(wide_merge.min_row + 1, wide_merge.max_row + 1):
                skip_rows.add(skip_r)
            continue

        kind = classify_row(row_values, is_title)

        if kind == 'blank':
            flush_table_buffer()
            continue

        if kind == 'title':
            flush_table_buffer()
            # Print just ONE value — merged cells fill every column with
            # the same repeated text, so non_empty[0] alone is the real title
            non_empty = [v for v in row_values if v is not None]
            if non_empty:
                parts.append(f"\n## {non_empty[0]}\n")
            continue

        if kind == 'key_value':
            flush_table_buffer()
            line = render_key_value_row(row_values)
            if line:
                parts.append(line)
            continue

        # table_row — accumulate into the buffer; flushed on blank/title/key_value
        table_buffer.append(row_values)

    flush_table_buffer()

    return '\n\n'.join(p for p in parts if p and p.strip())


def extract_xlsx_to_markdown(file_path: str) -> dict:
    path = Path(file_path)
    warnings = []
    sheet_mds = []

    try:
        wb = openpyxl.load_workbook(file_path, data_only=True)

        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            try:
                md = sheet_to_markdown(ws)
                if md.strip():
                    sheet_mds.append(f"### Sheet: {sheet_name}\n\n{md}")
                else:
                    warnings.append(f"Sheet '{sheet_name}': no content extracted")
            except Exception as e:
                warnings.append(f"Sheet '{sheet_name}': {e}")

        full_md = '\n\n---\n\n'.join(sheet_mds)
        quality = ('failed'  if not full_md.strip() else
                   'partial' if warnings else
                   'full')

        return {
            'file'    : path.name,
            'sheets'  : wb.sheetnames,
            'markdown': full_md,
            'quality' : quality,
            'warnings': warnings
        }

    except Exception as e:
        return {
            'file'    : path.name,
            'sheets'  : [],
            'markdown': '',
            'quality' : 'failed',
            'warnings': [f'Cannot open file: {e}']
        }


if __name__ == '__main__':
    result = extract_xlsx_to_markdown('/mnt/user-data/uploads/DSEC-AD-1X_Final_Well_Status_Report.xlsx')
    print(f"Quality : {result['quality']}")
    if result['warnings']:
        print(f"Warnings: {result['warnings']}")
    print()
    print(result['markdown'])
